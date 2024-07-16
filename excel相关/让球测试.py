import requests
import pandas as pd
from bs4 import BeautifulSoup
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font

# 获取用户输入的开始日期和结束日期
start_date_input = input("请输入开始日期 (格式: YYYY-MM-DD): ")
end_date_input = input("请输入结束日期 (格式: YYYY-MM-DD): ")

# 验证日期格式
try:
    start_date = datetime.datetime.strptime(start_date_input, "%Y-%m-%d")
    end_date = datetime.datetime.strptime(end_date_input, "%Y-%m-%d")
except ValueError:
    print("日期格式错误，请输入有效的日期 (格式: YYYY-MM-DD)")
    exit()

# 确保开始日期早于结束日期
if start_date > end_date:
    print("开始日期不能晚于结束日期")
    exit()

# 循环抓取日期范围内的数据
data_frames = []
current_date = start_date
while current_date <= end_date:
    formatted_date = current_date.strftime("%Y-%m-%d")
    url = f'https://trade.500.com/jczq/?playid=269&g=2&date={formatted_date}'

    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36"
    }

    response = requests.get(url, headers=headers)
    content = BeautifulSoup(response.text, 'html.parser')
    directory = content.find('div', class_='bet-main')

    if directory is None:
        print(f"未找到 {formatted_date} 的数据，请检查输入的日期是否正确或网站结构是否发生变化。")
        current_date += datetime.timedelta(days=1)
        continue

    # 编号
    time = [th.text.strip() for th in directory.find_all('td', 'td-no')]

    # 赛事
    events = [th.text.strip() for th in directory.find_all('td', 'td-evt')]

    # 开赛时间
    startTime = [th.text.strip() for th in directory.find_all('td', 'td-endtime')]

    # 主队
    homeTeam = [th.text.strip() for th in directory.find_all('a', 'team-l')]

    # vs
    vss = [th.text.strip() for th in directory.find_all('i', 'team-bf')]

    # 客队
    visitors = [th.text.strip() for th in directory.find_all('a', 'team-r')]

    # 让球
    handicapTheBall = [th.text.strip() for th in directory.find_all('p', 'itm-rangA1')]

    # 提取 betbtn-row itm-rangB1 中的胜平负数据
    def extract_data(divs, data_value, class_value=None):
        result = []
        for div in divs:
            if class_value:
                p_tag = div.find('p', {'data-value': data_value, 'class': class_value})
            else:
                p_tag = div.find('p', {'data-value': data_value})
            if p_tag:
                result.append(p_tag.text.strip())
            else:
                result.append(None)
        return result

    divs = directory.find_all('div', class_='betbtn-row itm-rangB1')

    # 胜
    victory = extract_data(divs, '3')
    victory_ok = [div.find('p', {'data-value': '3', 'class': 'betbtn betbtn-ok'}) is not None for div in divs]

    # 平
    flat = extract_data(divs, '1')
    flat_ok = [div.find('p', {'data-value': '1', 'class': 'betbtn betbtn-ok'}) is not None for div in divs]

    # 负
    negative = extract_data(divs, '0')
    negative_ok = [div.find('p', {'data-value': '0', 'class': 'betbtn betbtn-ok'}) is not None for div in divs]

    # 主让球
    theHomeTeamHandicaps = [th.text.strip() for th in directory.find_all('p', 'itm-rangA2')]

    # 提取 betbtn-row itm-rangB2 中的胜平负数据
    divse = directory.find_all('div', class_='betbtn-row itm-rangB2')

    # 主让胜
    victorys = extract_data(divse, '3')
    victorys_ok = [div.find('p', {'data-value': '3', 'class': 'betbtn betbtn-ok'}) is not None for div in divse]

    # 平
    flats = extract_data(divse, '1')
    flats_ok = [div.find('p', {'data-value': '1', 'class': 'betbtn betbtn-ok'}) is not None for div in divse]

    # 负
    negatives = extract_data(divse, '0')
    negatives_ok = [div.find('p', {'data-value': '0', 'class': 'betbtn betbtn-ok'}) is not None for div in divse]

    # 确保所有列具有相同的长度
    max_length = max(len(time), len(events), len(startTime), len(homeTeam), len(vss), len(visitors), len(handicapTheBall), len(victory), len(flat), len(negative), len(theHomeTeamHandicaps), len(victorys), len(flats), len(negatives))

    def pad_list(lst, length):
        return lst + [None] * (length - len(lst))

    time = pad_list(time, max_length)
    events = pad_list(events, max_length)
    startTime = pad_list(startTime, max_length)
    homeTeam = pad_list(homeTeam, max_length)
    vss = pad_list(vss, max_length)
    visitors = pad_list(visitors, max_length)
    handicapTheBall = pad_list(handicapTheBall, max_length)
    victory = pad_list(victory, max_length)
    victory_ok = pad_list(victory_ok, max_length)
    flat = pad_list(flat, max_length)
    flat_ok = pad_list(flat_ok, max_length)
    negative = pad_list(negative, max_length)
    negative_ok = pad_list(negative_ok, max_length)
    theHomeTeamHandicaps = pad_list(theHomeTeamHandicaps, max_length)
    victorys = pad_list(victorys, max_length)
    victorys_ok = pad_list(victorys_ok, max_length)
    flats = pad_list(flats, max_length)
    flats_ok = pad_list(flats_ok, max_length)
    negatives = pad_list(negatives, max_length)
    negatives_ok = pad_list(negatives_ok, max_length)

    # 创建DataFrame
    df = pd.DataFrame({
        '日期': [formatted_date] * max_length,
        '编号': time,
        '赛事': events,
        '开赛时间': startTime,
        '主队': homeTeam,
        'vs': vss,
        '客队': visitors,
        '让球': handicapTheBall,
        '胜': victory,
        '平': flat,
        '负': negative,
        '主队让球': theHomeTeamHandicaps,
        '让胜': victorys,
        '让平': flats,
        '让负': negatives
    })

    data_frames.append(df)
    current_date += datetime.timedelta(days=1)

# 合并所有 DataFrame
final_df = pd.concat(data_frames, ignore_index=True)

# 保存带有样式的 DataFrame 到 Excel 文件
output_file = f'{start_date_input}_to_{end_date_input}.xlsx'
final_df.to_excel(output_file, engine='openpyxl', index=False)

# 打开刚刚保存的 Excel 文件并应用样式
wb = load_workbook(output_file)
ws = wb.active

# 应用红色字体样式
red_font = Font(color="FF0000")

# 遍历表格，根据条件应用红色字体样式
# for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
#     for cell, ok_value in zip(row[8:], victory_ok + flat_ok + negative_ok + victorys_ok + flats_ok + negatives_ok):
#         if ok_value:
#             cell.font = red_font

# 保存应用了样式的 Excel 文件
wb.save(output_file)

print(f"已生成 {output_file}，并标注了红色字体。")


