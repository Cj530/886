import requests
import pandas as pd
from bs4 import BeautifulSoup
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

# 获取用户输入的开始日期和结束日期
start_date_input = input("请输入开始日期 (格式: YYYY-MM-DD): ")
end_date_input = input("请输入结束日期 (格式: YYYY-MM-DD): ")

# 验证日期格式
try:
    start_date = datetime.datetime.strptime(start_date_input, "%Y-%m-%d")
    end_date = datetime.datetime.strptime(end_date_input, "%Y-%m-%d")
except ValueError:
    print("日期格式错误，请输入有效的日期 (格式: YYYY-MM-DD)")
    exit()

# 确保开始日期早于等于结束日期
if start_date > end_date:
    print("开始日期不能晚于结束日期")
    exit()

# 创建一个空的DataFrame列表，用于存储每天的数据
data_frames = []

# 循环抓取日期范围内的数据
current_date = start_date
while current_date <= end_date:
    formatted_date = current_date.strftime("%Y-%m-%d")
    url = f'https://trade.500.com/jczq/?playid=271&g=2&date={formatted_date}'

    headers = {
        "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/109.0.0.0 Safari/537.36"
    }

    response = requests.get(url, headers=headers)
    content = BeautifulSoup(response.text, 'html.parser')
    directory = content.find('div', class_='bet-main')

    if directory is None:
        print(f"未找到 {formatted_date} 的数据，请检查输入的日期是否正确或网站结构是否发生变化。")
        current_date += datetime.timedelta(days=1)
        continue

    # 提取数据
    time = [th.text.strip() for th in directory.find_all('td', 'td-no')]
    events = [th.text.strip() for th in directory.find_all('td', 'td-evt')]
    startTime = [th.text.strip() for th in directory.find_all('td', 'td-endtime')]
    homeTeam = [th.text.strip() for th in directory.find_all('a', 'team-l')]
    vss = [th.text.strip() for th in directory.find_all('i', 'team-bf')]
    visitors = [th.text.strip() for th in directory.find_all('a', 'team-r')]
    Balls = [th.text.strip() for th in directory.find_all(attrs={'data-value':'3-3'})]
    
    Bonea = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'1:0'})]
    Boneb = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'2:0'})]
    Bonec = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'2:1'})]
    Boned = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'3:0'})]
    Bonee = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'3:1'})]
    Bonef = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'3:2'})]
    Boneg = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'4:0'})]
    Boneh = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'4:1'})]
    Bonei = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'4:2'})]
    Bonej = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'5:0'})]
    Bonek = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'5:1'})]
    Bonel = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'5:2'})]
    Bonem = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'胜其它'})]
    Bonen = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'0:0'})]
    Boneo = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'1:1'})]
    Bonep = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'2:2'})]
    Boneq = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'3:3'})]
    Boner = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'平其它'})]
    Bones = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'0:1'})]
    Bonet = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'0:2'})]
    Boneu = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'1:2'})]
    Bonev = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'0:3'})]
    Bonew = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'1:3'})]
    Bonex = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'2:3'})]
    Boney = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'0:4'})]
    Bonez = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'1:4'})]
    Boneza = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'2:4'})]
    Bonezb = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'0:5'})]
    Bonezc = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'1:5'})]
    Bonezd = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'2:5'})]
    Boneze = [p.find('i').text.strip() for p in directory.find_all(attrs={'data-value':'负其它'})]
 

    # 确保所有列具有相同的长度
    max_length = max(len(time), len(events), len(startTime), len(homeTeam), len(vss), len(visitors), len(Balls), len(Bonea),len(Boneb),len(Bonec),len(Boned),len(Bonee),len(Bonef),len(Boneg),len(Boneh),len(Bonei),len(Bonej),len(Bonek),len(Bonel),len(Bonem),len(Bonen),len(Boneo),len(Bonep),len(Boneq),len(Boner),len(Bones),len(Bonet),len(Boneu),len(Bonev),len(Bonew),len(Bonex),len(Boney),len(Bonez),len(Boneza),len(Bonezb),len(Bonezc),len(Bonee),len(Bonef),)

    def pad_list(lst, length):
        return lst + [None] * (length - len(lst))

    time = pad_list(time, max_length)
    events = pad_list(events, max_length)
    startTime = pad_list(startTime, max_length)
    homeTeam = pad_list(homeTeam, max_length)
    vss = pad_list(vss, max_length)
    visitors = pad_list(visitors, max_length)
    Balls = pad_list(Balls, max_length)
    Bonea = pad_list(Bonea, max_length)
    Boneb = pad_list(Boneb, max_length)
    Bonec = pad_list(Bonec, max_length)
    Boned = pad_list(Boned, max_length)
    Bonee = pad_list(Bonee, max_length)
    Bonef = pad_list(Bonef, max_length)
    Boneg = pad_list(Boneg, max_length)
    Boneh = pad_list(Boneh, max_length)
    Bonei = pad_list(Bonei, max_length)
    Bonej = pad_list(Bonej, max_length)
    Bonek = pad_list(Bonek, max_length)
    Bonel = pad_list(Bonel, max_length)
    Bonem = pad_list(Bonem, max_length)
    Bonen = pad_list(Bonen, max_length)
    Boneo = pad_list(Boneo, max_length)
    Bonep = pad_list(Bonep, max_length)
    Boneq = pad_list(Boneq, max_length)
    Boner = pad_list(Boner, max_length)
    Bones = pad_list(Bones, max_length)
    Bonet = pad_list(Bonet, max_length)
    Boneu = pad_list(Boneu, max_length)
    Bonev = pad_list(Bonev, max_length)
    Bonew = pad_list(Bonew, max_length)
    Bonex = pad_list(Bonex, max_length)
    Boney = pad_list(Boney, max_length)
    Bonez = pad_list(Bonez, max_length)
    Boneza = pad_list(Boneza, max_length)
    Bonezb = pad_list(Bonezb, max_length)
    Bonezc = pad_list(Bonezc, max_length)
    Bonezd = pad_list(Bonezd, max_length)
    Boneze = pad_list(Boneze, max_length)



    
    # 创建DataFrame
    df = pd.DataFrame({
        '日期': [formatted_date] * max_length,
        '编号': time,
        '赛事': events,
        '开赛时间': startTime,
        '主队': homeTeam,
        'vs': vss,
        '客队': visitors,
        '1:0': Bonea,
        '2:0': Boneb,
        '2:1': Bonec,
        '3:0': Boned,
        '3:1': Bonee,
        '3:2': Bonef,
        '4:0': Boneg,
        '4:1': Boneh,
        '4:2': Bonei,
        '5:0': Bonej,
        '5:1': Bonek,
        '5:2': Bonel,
        '胜其它': Bonem,
        '0:0': Bonen,
        '1:1':Boneo,
        '2:2':Bonep,
        '3:3':Boneq,
        '平其它':Boner,
        '0:1':Bones,
        '0:2':Bonet,
        '1:2':Boneu,
        '0:3':Bonev,
        '1:3':Bonew,
        '2:3':Bonex,
        '0:4':Boney,
        '1:4':Bonez,
        '2:4':Boneza,
        '0:5':Bonezb,
        '1:5':Bonezc,
        '2:5':Bonezd,
        '负其它':Boneze,

    })

    
    data_frames.append(df)
    current_date += datetime.timedelta(days=1)
    

# 合并所有 DataFrame
final_df = pd.concat(data_frames, ignore_index=True)

# 保存到 Excel 文件
output_file = f'{start_date_input}_to_{end_date_input}.xlsx'
final_df.to_excel(output_file, engine='openpyxl', index=False)

# 打开并设置样式
wb = load_workbook(output_file)
ws = wb.active

# 设置红色字体样式
red_font = Font(color="FF0000")

# 遍历表格并应用红色字体样式
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=15):
    for cell in row:
        if cell.value:
            cell.font = red_font

# 保存 Excel 文件
wb.save(output_file)

print(f"已生成 {output_file}，并标注了红色字体。")
